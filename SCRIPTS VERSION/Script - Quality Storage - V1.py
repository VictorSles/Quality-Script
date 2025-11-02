import requests
import threading
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import pandas as pd
import os
import warnings
from datetime import datetime
import subprocess
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

warnings.filterwarnings("ignore", message="Unverified HTTPS request")

# =======================
# CONFIGURAÇÕES
# =======================
API_URL_BASE = "https://MAN-prd.jemsms.corp.jabil.org/"
USER = r"jabil\svchua_jesmapistg"
PASSWORD = "qKzla3oBDA51Ecq=+B2_z"

# Caminho de rede onde serão salvos os arquivos
NETWORK_PATH = r"\\172.24.80.153\axi-aoi\QUALIDADE\INJET FOLDER LOG"
os.makedirs(NETWORK_PATH, exist_ok=True)

# Log diário, nunca sobrescreve
LOG_FILE = os.path.join(NETWORK_PATH, f"wip_log_{datetime.now():%Y%m%d}.txt")
XLSX_FILE = os.path.join(NETWORK_PATH, f"wip_log_exportado_{datetime.now():%Y%m%d}.xlsx")

LINHAS = [
    "Rio Jari", "Rio Jutaí", "Rio Negro",
    "Rio Japurá", "Rio Juruá", "Rio Xingu", "RioTefé"
]
LADOS = ["TOP", "BOT"]

_cached_token = None

# =======================
# FUNÇÕES DE REDE
# =======================
def ensure_network_connection():
    """Verifica se a pasta de rede está acessível; caso não, pede login e autentica."""
    if os.path.exists(NETWORK_PATH):
        return True

    messagebox.showwarning(
        "Conexão de Rede",
        f"A pasta de rede não está acessível:\n{NETWORK_PATH}\n\nSerá necessário autenticar..."
    )

    username = simpledialog.askstring("Login de Rede", "Usuário (ex: jabil\\seu_usuario):")
    password = simpledialog.askstring("Senha de Rede", "Senha:", show="*")
    if not username or not password:
        messagebox.showerror("Erro", "Credenciais de rede não fornecidas.")
        return False

    try:
        cmd = f'net use "{NETWORK_PATH}" /user:{username} "{password}"'
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
        if result.returncode == 0:
            messagebox.showinfo("Conectado", f"Conectado com sucesso à pasta de rede:\n{NETWORK_PATH}")
            return True
        else:
            messagebox.showerror("Erro", f"Falha na autenticação:\n{result.stderr}")
            return False
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao conectar à rede:\n{e}")
        return False

# =======================
# AUTENTICAÇÃO API
# =======================
def get_token():
    global _cached_token
    if _cached_token:
        return _cached_token

    url = f"{API_URL_BASE}api-external-api/api/user/adsignin"
    form_data = {"name": USER, "password": PASSWORD}
    try:
        resp = requests.post(url, data=form_data, verify=False, timeout=15)
        resp.raise_for_status()
        _cached_token = resp.text.strip()
        print("✅ Token obtido")
        return _cached_token
    except Exception as e:
        print("Erro get_token:", e)
        return None

def build_session():
    token = get_token()
    if not token:
        return None
    try:
        if "=" in token:
            cookie_name, cookie_value = token.split("=", 1)
        else:
            cookie_name, cookie_value = "AuthToken", token
    except Exception:
        print("Token em formato inesperado:", token)
        return None
    session = requests.Session()
    session.verify = False
    session.cookies.set(cookie_name.strip(), cookie_value.strip(";"))
    session.headers.update({"Cache-Control": "no-cache"})
    return session

# =======================
# API WIP INFO
# =======================
def get_wip_info(session, serial):
    try:
        url = f"{API_URL_BASE}api-external-api/api/Wips/getWipInformationBySerialNumber"
        params = {"SiteName": "MANAUS", "CustomerName": "SAMSUNG", "SerialNumber": serial}
        r = session.get(url, params=params, timeout=15)
        if r.status_code != 200:
            print(f"get_wip_info: status {r.status_code} - {r.text[:200]}")
            return None
        data = r.json()
        return data[0] if isinstance(data, list) and data else None
    except Exception as e:
        print("Erro get_wip_info:", e)
        return None

# =======================
# EXTRAÇÃO & LOG
# =======================
def extract_info(wip_data):
    return {
        "Serial": wip_data.get("SerialNumber", "N/A"),
        "Modelo": wip_data.get("MaterialName", "N/A"),
        "FERT": wip_data.get("MaterialName", "N/A"),
        "Descricao": wip_data.get("AssemblyDescription", "N/A"),
        "Revision": wip_data.get("AssemblyRevision", "N/A"),
        "Versao": wip_data.get("AssemblyVersion", "N/A"),
        "Ordem": wip_data.get("PlannedOrderNumber", "N/A"),
        "WipStatus": wip_data.get("WipStatus", "N/A"),
        "Criado": (wip_data.get("WipCreationDate") or "")[:10],
    }

def format_info_line(info: dict, linha: str, lado: str):
    agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    parts = [f"{k}: {v}" for k, v in info.items()]
    parts.append(f"Linha: {linha}")
    parts.append(f"Lado: {lado}")
    parts.append(f"DataHoraProcessamento: {agora}")
    return " | ".join(parts)

def append_to_log(wip_data, linha: str, lado: str, session):
    serial = wip_data.get("SerialNumber")
    if serial is None:
        return 0

    if not ensure_network_connection():
        return 0

    count = 0
    try:
        os.makedirs(NETWORK_PATH, exist_ok=True)
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            info = extract_info(wip_data)
            linha_fmt = format_info_line(info, linha, lado)
            f.write("\n" + linha_fmt.strip() + "\n")
            count += 1

            panel = wip_data.get("Panel")
            if panel and "PanelWips" in panel:
                for pw in panel["PanelWips"]:
                    pw_serial = pw.get("SerialNumber")
                    if pw_serial and pw_serial != serial:
                        wip_link = get_wip_info(session, pw_serial)
                        if wip_link:
                            info_link = extract_info(wip_link)
                            linha_fmt2 = format_info_line(info_link, linha, lado)
                            f.write("\n" + linha_fmt2.strip() + "\n")
                            count += 1
        return count
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao escrever log:\n{e}")
        return 0

# =======================
# EXPORTAÇÃO COM GRÁFICO
# =======================
def export_to_excel():
    if not os.path.exists(LOG_FILE):
        messagebox.showerror("Erro", f"Arquivo de log não encontrado:\n{LOG_FILE}")
        return
    try:
        # Lê o log
        with open(LOG_FILE, "r", encoding="utf-8") as f:
            linhas = [l.strip() for l in f if l.strip()]

        registros = []
        for linha in linhas:
            d = {}
            for part in linha.split(" | "):
                if ": " in part:
                    k, v = part.split(": ", 1)
                    d[k.strip()] = v.strip()
            registros.append(d)

        df = pd.DataFrame(registros)
        df = df.fillna("N/A")

        # Salva Excel
        df.to_excel(XLSX_FILE, index=False)

        # Criação do gráfico
        wb = load_workbook(XLSX_FILE)
        ws = wb.active
        ws.title = "WIP Log"

        # Resumo por FERT
        resumo = df["FERT"].value_counts().reset_index()
        resumo.columns = ["FERT", "Quantidade"]

        ws2 = wb.create_sheet("Resumo")
        ws2.append(["FERT", "Quantidade"])
        for row in resumo.itertuples(index=False):
            ws2.append(row)

        chart = BarChart()
        chart.title = "Quantidade de Placas por Modelo (FERT)"
        chart.x_axis.title = "FERT (Modelo)"
        chart.y_axis.title = "Quantidade"

        data = Reference(ws2, min_col=2, min_row=1, max_row=len(resumo) + 1)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=len(resumo) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 20

        ws2.add_chart(chart, "E2")
        wb.save(XLSX_FILE)
        wb.close()

        messagebox.showinfo("Exportar Excel", f"✅ Exportado com gráfico!\n{XLSX_FILE}")

    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao exportar para Excel:\n{e}")

# =======================
# LIMPAR LOG (ARQUIVAR)
# =======================
def clear_log():
    if not ensure_network_connection():
        return
    try:
        if os.path.exists(LOG_FILE):
            backup_path = LOG_FILE.replace(".txt", f"_{datetime.now():%H%M%S}_backup.txt")
            os.rename(LOG_FILE, backup_path)
            messagebox.showinfo("Limpar Log", f"Arquivo atual arquivado como:\n{os.path.basename(backup_path)}")
        else:
            messagebox.showinfo("Limpar Log", "Nenhum arquivo de log encontrado para arquivar.")
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao arquivar log:\n{e}")

# =======================
# INTERFACE TKINTER
# =======================
class WipApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("🧠 Jabil MES - WIP Info Extractor (Rede)")
        self.geometry("600x500")
        self.resizable(False, False)

        self.session = build_session()
        if not self.session:
            messagebox.showerror("Erro", "Falha na autenticação do MES.")
            self.destroy()
            return

        self.processados = 0
        self.create_widgets()

    def create_widgets(self):
        pad = 10
        ttk.Label(self, text="Serial:").pack(pady=(pad, 2))
        self.serial_entry = ttk.Entry(self, width=45)
        self.serial_entry.pack()
        self.serial_entry.bind("<Return>", lambda e: self.on_process())

        ttk.Label(self, text="Linha:").pack(pady=(pad, 2))
        self.linha_var = tk.StringVar(value=LINHAS[0])
        ttk.Combobox(self, textvariable=self.linha_var, state="readonly", values=LINHAS).pack()

        ttk.Label(self, text="Lado da Placa:").pack(pady=(pad, 2))
        self.lado_var = tk.StringVar(value=LADOS[0])
        ttk.Combobox(self, textvariable=self.lado_var, state="readonly", values=LADOS).pack()

        ttk.Button(self, text="Processar", command=self.on_process).pack(pady=(pad, 5))
        ttk.Button(self, text="📤 Exportar Excel", command=export_to_excel).pack(pady=5)
        ttk.Button(self, text="🗑️ Limpar Log", command=clear_log).pack(pady=5)

        self.progress = ttk.Progressbar(self, orient="horizontal", length=450, mode="determinate")
        self.progress.pack(pady=(pad, 5))
        self.status_label = ttk.Label(self, text="")
        self.status_label.pack(pady=5)

        self.result_text = tk.Text(self, height=10, width=70)
        self.result_text.pack(pady=5)

    def on_process(self):
        serial = self.serial_entry.get().strip()
        linha = self.linha_var.get().strip()
        lado = self.lado_var.get().strip()
        if not serial:
            messagebox.showwarning("Atenção", "Digite o número de série.")
            return
        threading.Thread(target=self.process_serial, args=(serial, linha, lado), daemon=True).start()

    def process_serial(self, serial, linha, lado):
        self.progress["value"] = 10
        self.status_label.config(text="Buscando dados...")
        self.update()

        wip = get_wip_info(self.session, serial)
        if not wip:
            self.status_label.config(text="❌ Serial não encontrado ou erro na API.")
            return

        self.progress["value"] = 50
        cont = append_to_log(wip, linha, lado, self.session)
        self.progress["value"] = 100

        if cont == 0:
            self.status_label.config(text="⚠️ Nenhum dado gravado (erro ou acesso negado).")
        else:
            self.processados += cont
            resumo = format_info_line(extract_info(wip), linha, lado)
            self.result_text.delete("1.0", tk.END)
            self.result_text.insert(tk.END, resumo + "\n")
            self.status_label.config(text=f"✅ {serial} processado. (+{cont-1} vinculados). Total: {self.processados}")
            messagebox.showinfo("Arquivo Gerado", f"Salvo em:\n{LOG_FILE}")

        self.progress["value"] = 0
        self.serial_entry.delete(0, tk.END)

# =======================
# EXECUÇÃO
# =======================
if __name__ == "__main__":
    app = WipApp()
    app.mainloop()
